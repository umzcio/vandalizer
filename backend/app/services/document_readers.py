"""Multi-format text extraction (PDF, DOCX, XLSX, HTML, etc.).

Ported from Flask app/utilities/document_readers.py.
All functions are synchronous — safe for Celery workers.
"""

import logging
import os
import re
from datetime import date, datetime, time

from markitdown import MarkItDown

logger = logging.getLogger(__name__)

MIN_PDF_TEXT_LENGTH = 100
MAX_XLSX_COMMENT_LEN = 500


def clean_markdown_nans(markdown_content: str) -> str:
    """Remove NaN values from markdown content."""
    cleaned = markdown_content.replace("| NaN |", "| |")
    cleaned = cleaned.replace("NaN", "")

    lines = cleaned.split("\n")
    filtered_lines = []
    for line in lines:
        if "|" in line:
            cells = [cell.strip() for cell in line.split("|")[1:-1]]
            if any(cell and cell != "---" for cell in cells) or all(
                cell in ["---", ""] for cell in cells
            ):
                filtered_lines.append(line)
        else:
            filtered_lines.append(line)

    return "\n".join(filtered_lines)


def convert_to_markdown(doc_path: str, keep_data_uris: bool = True) -> str:
    """Convert a document to Markdown format using MarkItDown."""
    md = MarkItDown(enable_plugins=False)
    result = md.convert(doc_path, keep_data_uris=keep_data_uris)
    return clean_markdown_nans(result.text_content)


def extract_text_from_pdf(pdf_path: str) -> str:
    """Extract text from a PDF using PyMuPDF.

    PyMuPDF preserves reading order in multi-column layouts and exposes
    form field values that PyPDF2 misses (NIH biosketches, NSF Current
    & Pending forms, etc. are common research-admin uploads).
    """
    import pymupdf

    chunks: list[str] = []
    with pymupdf.open(pdf_path) as doc:
        for page in doc:
            chunks.append(page.get_text("text"))

            field_lines: list[str] = []
            for widget in page.widgets() or []:
                value = (widget.field_value or "").strip()
                if not value:
                    continue
                label = (widget.field_label or widget.field_name or "").strip()
                field_lines.append(f"- {label}: {value}" if label else f"- {value}")
            if field_lines:
                chunks.append("[Form fields]\n" + "\n".join(field_lines))

    return "\n".join(c for c in chunks if c)


def ocr_extract_text_from_pdf(pdf_path: str, retries: int = 3) -> str:
    """Extract text from a PDF using the UIPDF OCR endpoint.

    Falls back gracefully if the OCR service is unavailable.
    """
    # OCR endpoint is stored in the database via admin config (SystemConfig)
    from pymongo import MongoClient
    from app.config import Settings
    from app.utils.encryption import decrypt_value
    settings = Settings()
    client = MongoClient(settings.mongo_host)
    db = client[settings.mongo_db]
    cfg = db.system_config.find_one({})
    ocr_endpoint = (cfg or {}).get("ocr_endpoint", "")
    raw_api_key = (cfg or {}).get("ocr_api_key", "")
    ocr_api_key = decrypt_value(raw_api_key)

    if not ocr_endpoint:
        logger.warning("OCR_ENDPOINT not configured — skipping OCR for %s", pdf_path)
        return ""

    # If decrypt_value returned the raw 'enc:' ciphertext, CONFIG_ENCRYPTION_KEY
    # is missing or wrong in this process (commonly the Celery worker env).
    if ocr_api_key.startswith("enc:"):
        logger.error(
            "OCR api key could not be decrypted — CONFIG_ENCRYPTION_KEY missing "
            "or mismatched in this worker. Fix the env var and restart Celery."
        )
        return ""

    logger.info(
        "Extracting text with OCR: endpoint=%s key_set=%s key_len=%d file=%s",
        ocr_endpoint, bool(ocr_api_key), len(ocr_api_key), pdf_path,
    )

    headers = {}
    if ocr_api_key:
        headers["Authorization"] = f"Bearer {ocr_api_key}"

    import httpx
    for attempt in range(retries):
        try:
            with httpx.Client(timeout=120.0) as client:
                with open(pdf_path, "rb") as f:
                    resp = client.post(
                        ocr_endpoint,
                        headers=headers,
                        files={"file": (os.path.basename(pdf_path), f, "application/pdf")},
                    )
                if resp.status_code == 200:
                    return resp.text
                logger.warning(
                    "OCR attempt %d returned HTTP %d from %s — body: %s",
                    attempt + 1, resp.status_code, ocr_endpoint, resp.text[:500],
                )
        except Exception as e:
            logger.warning("OCR attempt %d raised: %s", attempt + 1, e)

    logger.error("OCR failed after %d attempts for %s", retries, pdf_path)
    return ""


def _stringify_cell_value(value: object) -> str:
    """Render an openpyxl cell value as a plain display string."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        rounded = round(value, 4)
        text = f"{rounded:.4f}".rstrip("0").rstrip(".")
        return text or "0"
    return str(value)


def _format_xlsx_cell(value: object) -> str:
    """Render an openpyxl cell value as a pipe-table-safe string."""
    text = _stringify_cell_value(value)
    if not text:
        return ""
    return text.replace("\\", "\\\\").replace("|", r"\|").replace("\n", " ").strip()


def _evaluate_xlsx_formulas(xlsx_path: str) -> dict[tuple[str, str], object]:
    """Compute formula results with the `formulas` library.

    Returns a {(sheet_name_upper, coordinate_upper): value} map. Empty on
    failure — callers fall back to formula text.
    """
    try:
        import formulas
    except ImportError:
        return {}

    try:
        model = formulas.ExcelModel().loads(xlsx_path).finish()
        solution = model.calculate()
    except Exception as e:
        logger.warning("formulas evaluation failed for %s: %s", xlsx_path, e)
        return {}

    result: dict[tuple[str, str], object] = {}
    for key, cell in solution.items():
        # Keys look like: "'[file.xlsx]SHEETNAME'!A1" or with ranges "...!A1:B2"
        try:
            sheet_part, coord = key.split("!", 1)
            if ":" in coord:
                continue  # skip range entries
            sheet_name = sheet_part.split("]", 1)[1].rstrip("'").upper()
        except Exception:
            continue

        value = cell.value if hasattr(cell, "value") else cell
        try:
            if hasattr(value, "tolist"):
                unwrapped = value.tolist()
                while isinstance(unwrapped, list) and len(unwrapped) == 1:
                    unwrapped = unwrapped[0]
                value = unwrapped
        except Exception:
            pass

        # `formulas` returns its own error sentinels for #DIV/0 etc.
        type_name = type(value).__name__
        if type_name in {"XlError", "Empty"}:
            continue

        result[(sheet_name, coord.upper())] = value

    return result


def extract_text_from_xlsx(xlsx_path: str) -> str:
    """Extract every visible and hidden sheet from an .xlsx workbook.

    Why this exists: MarkItDown (pandas under the hood) silently drops
    cells that have a formula but no cached calculated value. Workbooks
    saved by Google Sheets or headless LibreOffice commonly land in that
    state, so entire budget columns come through as blank. This walker
    uses openpyxl directly, prefers cached values, evaluates missing
    formulas with the `formulas` library when available, and falls back
    to formula text as a last resort.
    """
    import openpyxl

    try:
        wb_values = openpyxl.load_workbook(xlsx_path, data_only=True)
        wb_formulas = openpyxl.load_workbook(xlsx_path, data_only=False)
    except Exception as e:
        logger.warning(
            "openpyxl failed on %s (%s) — falling back to MarkItDown", xlsx_path, e
        )
        return convert_to_markdown(xlsx_path, keep_data_uris=False)

    # Only run the formula engine if we actually need it (any uncached formulas).
    needs_eval = False
    for sheet_name in wb_values.sheetnames:
        ws_v = wb_values[sheet_name]
        ws_f = wb_formulas[sheet_name]
        for r in range(1, (ws_v.max_row or 0) + 1):
            for c in range(1, (ws_v.max_column or 0) + 1):
                if ws_v.cell(row=r, column=c).value is None:
                    fv = ws_f.cell(row=r, column=c).value
                    if isinstance(fv, str) and fv.startswith("="):
                        needs_eval = True
                        break
            if needs_eval:
                break
        if needs_eval:
            break

    computed = _evaluate_xlsx_formulas(xlsx_path) if needs_eval else {}

    out: list[str] = []

    for sheet_name in wb_values.sheetnames:
        ws_v = wb_values[sheet_name]
        ws_f = wb_formulas[sheet_name]

        max_row = ws_v.max_row or 0
        max_col = ws_v.max_column or 0
        if max_row == 0 or max_col == 0:
            continue

        sheet_key = sheet_name.upper()
        grid: list[list[object]] = []
        for r in range(1, max_row + 1):
            row: list[object] = []
            for c in range(1, max_col + 1):
                cell_v = ws_v.cell(row=r, column=c)
                cached = cell_v.value
                formula = ws_f.cell(row=r, column=c).value
                if cached is None and isinstance(formula, str) and formula.startswith("="):
                    coord = cell_v.coordinate.upper()
                    evaluated = computed.get((sheet_key, coord))
                    row.append(evaluated if evaluated is not None else formula)
                else:
                    row.append(cached)
            grid.append(row)

        kept_rows = [row for row in grid if any(v not in (None, "") for v in row)]
        if not kept_rows:
            continue

        keep_col = [
            any(row[c] not in (None, "") for row in kept_rows) for c in range(max_col)
        ]
        trimmed = [[row[c] for c, keep in enumerate(keep_col) if keep] for row in kept_rows]

        header = [f"## {sheet_name}"]
        if ws_v.sheet_state != "visible":
            header.append(f"_(sheet is {ws_v.sheet_state})_")
        out.append("\n".join(header))

        lines = ["| " + " | ".join(_format_xlsx_cell(v) for v in row) + " |" for row in trimmed]
        if len(lines) > 1:
            sep = "| " + " | ".join("---" for _ in trimmed[0]) + " |"
            lines.insert(1, sep)
        out.append("\n".join(lines))

        extras: list[str] = []
        merged = list(ws_v.merged_cells.ranges)
        if merged:
            extras.append(f"_Merged ranges: {', '.join(str(r) for r in merged)}_")

        hidden_rows = [
            r for r in range(1, max_row + 1) if ws_v.row_dimensions[r].hidden
        ]
        hidden_cols = [
            col for col, dim in ws_v.column_dimensions.items() if dim.hidden
        ]
        if hidden_rows:
            extras.append(f"_Hidden rows: {hidden_rows}_")
        if hidden_cols:
            extras.append(f"_Hidden columns: {hidden_cols}_")

        comments = []
        for row in ws_v.iter_rows():
            for cell in row:
                if cell.comment and cell.comment.text:
                    text = cell.comment.text.strip()
                    if len(text) > MAX_XLSX_COMMENT_LEN:
                        text = text[:MAX_XLSX_COMMENT_LEN] + "…"
                    comments.append((cell.coordinate, text))
        if comments:
            extras.append("_Cell comments:_")
            extras.extend(f"- {coord}: {text}" for coord, text in comments)

        if extras:
            out.append("\n".join(extras))

    defined = list(wb_formulas.defined_names)
    if defined:
        block = ["## Defined names"]
        for name in defined:
            try:
                value = wb_formulas.defined_names[name].value
            except Exception:
                value = "?"
            block.append(f"- {name}: {value}")
        out.append("\n".join(block))

    return "\n\n".join(out).strip()


def extract_sheet_json_from_xlsx(xlsx_path: str) -> dict:
    """Render an .xlsx workbook as JSON sheets for the document viewer.

    Mirrors extract_text_from_xlsx's evaluation strategy so the viewer
    and OCR agree on formula results: prefer Excel-cached values, then
    fall back to the `formulas` library, then to the formula text.
    """
    import openpyxl

    wb_values = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_formulas = openpyxl.load_workbook(xlsx_path, data_only=False)

    needs_eval = False
    for sheet_name in wb_values.sheetnames:
        ws_v = wb_values[sheet_name]
        ws_f = wb_formulas[sheet_name]
        for r in range(1, (ws_v.max_row or 0) + 1):
            for c in range(1, (ws_v.max_column or 0) + 1):
                if ws_v.cell(row=r, column=c).value is None:
                    fv = ws_f.cell(row=r, column=c).value
                    if isinstance(fv, str) and fv.startswith("="):
                        needs_eval = True
                        break
            if needs_eval:
                break
        if needs_eval:
            break

    computed = _evaluate_xlsx_formulas(xlsx_path) if needs_eval else {}

    sheets: list[dict] = []
    for sheet_name in wb_values.sheetnames:
        ws_v = wb_values[sheet_name]
        ws_f = wb_formulas[sheet_name]

        max_row = ws_v.max_row or 0
        max_col = ws_v.max_column or 0
        if max_row == 0 or max_col == 0:
            sheets.append({"name": sheet_name, "headers": [], "rows": [], "hidden": ws_v.sheet_state != "visible"})
            continue

        sheet_key = sheet_name.upper()
        grid: list[list[str]] = []
        for r in range(1, max_row + 1):
            row: list[str] = []
            for c in range(1, max_col + 1):
                cell_v = ws_v.cell(row=r, column=c)
                cached = cell_v.value
                formula = ws_f.cell(row=r, column=c).value
                if cached is None and isinstance(formula, str) and formula.startswith("="):
                    coord = cell_v.coordinate.upper()
                    evaluated = computed.get((sheet_key, coord))
                    row.append(_stringify_cell_value(evaluated if evaluated is not None else formula))
                else:
                    row.append(_stringify_cell_value(cached))
            grid.append(row)

        headers = grid[0] if grid else []
        rows = grid[1:] if len(grid) > 1 else []
        sheets.append({
            "name": sheet_name,
            "headers": headers,
            "rows": rows,
            "hidden": ws_v.sheet_state != "visible",
        })

    return {"sheets": sheets}


_DOCX_W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


def _docx_text_of(element) -> str:
    """Concatenate all <w:t> and <w:delText> descendants of a DOCX element."""
    parts = []
    for tag in ("t", "delText"):
        for node in element.iter(f"{{{_DOCX_W_NS}}}{tag}"):
            if node.text:
                parts.append(node.text)
    return "".join(parts).strip()


def extract_docx_extras(docx_path: str) -> str:
    """Pull comments and tracked changes from a .docx file.

    Returns a markdown block to append after the body, or "" if there's
    nothing notable. Research admins live in Word comments during
    proposal review, and pypandoc/MarkItDown both drop them silently.
    """
    import defusedxml.ElementTree as ET
    import zipfile

    try:
        zf = zipfile.ZipFile(docx_path)
    except (zipfile.BadZipFile, FileNotFoundError):
        return ""

    sections: list[str] = []

    with zf:
        names = set(zf.namelist())

        if "word/comments.xml" in names:
            try:
                tree = ET.fromstring(zf.read("word/comments.xml"))
            except ET.ParseError:
                tree = None
            if tree is not None:
                lines = []
                for c in tree.findall(f"{{{_DOCX_W_NS}}}comment"):
                    author = c.attrib.get(f"{{{_DOCX_W_NS}}}author", "Unknown")
                    date = c.attrib.get(f"{{{_DOCX_W_NS}}}date", "")
                    text = _docx_text_of(c)
                    if not text:
                        continue
                    header = f"- **{author}**"
                    if date:
                        header += f" ({date})"
                    lines.append(f"{header}: {text}")
                if lines:
                    sections.append("## Comments\n" + "\n".join(lines))

        if "word/document.xml" in names:
            try:
                doc_tree = ET.fromstring(zf.read("word/document.xml"))
            except ET.ParseError:
                doc_tree = None
            if doc_tree is not None:
                changes: list[str] = []
                for kind, label in (("ins", "Inserted"), ("del", "Deleted")):
                    for el in doc_tree.iter(f"{{{_DOCX_W_NS}}}{kind}"):
                        text = _docx_text_of(el)
                        if not text:
                            continue
                        author = el.attrib.get(f"{{{_DOCX_W_NS}}}author", "Unknown")
                        date = el.attrib.get(f"{{{_DOCX_W_NS}}}date", "")
                        suffix = f" ({date})" if date else ""
                        changes.append(f"- **{label}** by {author}{suffix}: {text}")
                if changes:
                    sections.append("## Tracked changes\n" + "\n".join(changes))

    return "\n\n".join(sections)


def remove_images_from_markdown(markdown_text: str) -> str:
    """Remove all image references and their size attributes from markdown text."""
    text = re.sub(r"!\[([^\]]*)\]\([^)]*\)", "", markdown_text)
    text = re.sub(r"!\[([^\]]*)\]\[[^\]]*\]", "", text)
    text = re.sub(r'\{[^}]*(?:width|height)\s*=\s*"[^"]*"[^}]*\}', "", text)
    text = re.sub(r'\{[^{}]*="[^"]*"[^{}]*\}', "", text)
    text = re.sub(r"^\s*\[[^\]]+\]:\s*[^\s]+.*$", "", text, flags=re.MULTILINE)
    text = re.sub(r"\n\s*\n\s*\n", "\n\n", text)
    text = re.sub(r"^\s+$", "", text, flags=re.MULTILINE)
    return text.strip()


def extract_text_from_file(file_path: str, file_extension: str) -> str:
    """Extract text from a file based on its extension.

    This is the primary entry point used by document_tasks.
    """
    file_extension = file_extension.lower().lstrip(".")

    try:
        if file_extension == "pdf":
            # Prefer OCR when available — it handles scanned pages,
            # complex layouts, and image-heavy PDFs far better than PyPDF2.
            ocr_text = ocr_extract_text_from_pdf(file_path)
            if ocr_text and len(ocr_text.strip()) >= MIN_PDF_TEXT_LENGTH:
                return ocr_text
            # Fall back to PyMuPDF if OCR unavailable or returned nothing.
            logger.info("OCR returned %d chars, falling back to PyMuPDF", len(ocr_text))
            text = extract_text_from_pdf(file_path)
            return text

        elif file_extension in ("html", "htm"):
            return convert_to_markdown(file_path, keep_data_uris=False)

        elif file_extension in ("txt", "md", "csv", "json", "xml", "log"):
            with open(file_path, encoding="utf-8") as f:
                return f.read()

        elif file_extension == "xlsx":
            return extract_text_from_xlsx(file_path)

        elif file_extension == "docx":
            body = convert_to_markdown(file_path, keep_data_uris=False)
            extras = extract_docx_extras(file_path)
            return (body.rstrip() + "\n\n" + extras) if extras else body

        elif file_extension in ("doc", "xls", "pptx", "ppt"):
            return convert_to_markdown(file_path, keep_data_uris=False)

        elif file_extension in ("py", "js", "java", "cpp", "c", "h", "css", "sql"):
            with open(file_path, encoding="utf-8") as f:
                return f.read()

        else:
            try:
                return convert_to_markdown(file_path, keep_data_uris=False)
            except Exception:
                try:
                    with open(file_path, encoding="utf-8") as f:
                        return f.read()
                except Exception:
                    with open(file_path, encoding="latin-1") as f:
                        return f.read()

    except Exception as e:
        logger.error("Error extracting text from %s: %s", file_path, e)
        return f"[Error extracting content: {e!s}]"
